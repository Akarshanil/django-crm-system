from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib import messages
from django.http import HttpResponse, FileResponse
from django.db.models import Q
from django.core.paginator import Paginator
from .models import Customer, UserProfile
from .forms import CustomerForm, BulkUploadForm, UserRegistrationForm, UserEditForm, UserProfileForm
import openpyxl
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from datetime import datetime


# Dashboard View
@login_required(login_url='login')
def dashboard(request):
    total_customers = Customer.objects.count()
    total_users = User.objects.count()
    recent_customers = Customer.objects.all()[:5]
    
    context = {
        'total_customers': total_customers,
        'total_users': total_users,
        'recent_customers': recent_customers,
    }
    return render(request, 'customers/dashboard.html', context)


# Customer List View
@login_required(login_url='login')
def customer_list(request):
    search_query = request.GET.get('search', '')
    
    if search_query:
        customers = Customer.objects.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(company__icontains=search_query)
        )
    else:
        customers = Customer.objects.all()
    
    paginator = Paginator(customers, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    return render(request, 'customers/customer_list.html', context)


# Customer Add View
@login_required(login_url='login')
def customer_add(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST, request.FILES)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.created_by = request.user
            customer.save()
            messages.success(request, 'Customer added successfully!')
            return redirect('customer_list')
    else:
        form = CustomerForm()
    
    context = {'form': form, 'title': 'Add Customer'}
    return render(request, 'customers/customer_form.html', context)


# Customer Edit View
@login_required(login_url='login')
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, request.FILES, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, 'Customer updated successfully!')
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
    
    context = {'form': form, 'title': 'Edit Customer', 'customer': customer}
    return render(request, 'customers/customer_form.html', context)


# Customer Detail View
@login_required(login_url='login')
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    context = {'customer': customer}
    return render(request, 'customers/customer_detail.html', context)


# Customer Delete View
@login_required(login_url='login')
def customer_delete(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        customer.delete()
        messages.success(request, 'Customer deleted successfully!')
        return redirect('customer_list')
    
    context = {'customer': customer}
    return render(request, 'customers/customer_confirm_delete.html', context)


# Bulk Upload View
@login_required(login_url='login')
def bulk_upload(request):
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                success_count = 0
                error_count = 0
                errors = []
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        if not row[0] or not row[1] or not row[2]:  # Check required fields
                            continue
                            
                        customer = Customer(
                            first_name=row[0],
                            last_name=row[1],
                            email=row[2],
                            phone=row[3] if len(row) > 3 and row[3] else '',
                            address=row[4] if len(row) > 4 and row[4] else '',
                            city=row[5] if len(row) > 5 and row[5] else '',
                            state=row[6] if len(row) > 6 and row[6] else '',
                            country=row[7] if len(row) > 7 and row[7] else '',
                            postal_code=row[8] if len(row) > 8 and row[8] else '',
                            company=row[9] if len(row) > 9 and row[9] else '',
                            notes=row[10] if len(row) > 10 and row[10] else '',
                            created_by=request.user
                        )
                        customer.save()
                        success_count += 1
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
                
                if success_count > 0:
                    messages.success(request, f'Successfully imported {success_count} customers!')
                if error_count > 0:
                    messages.warning(request, f'Failed to import {error_count} customers. Errors: {", ".join(errors[:5])}')
                
                return redirect('customer_list')
                
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    else:
        form = BulkUploadForm()
    
    context = {'form': form}
    return render(request, 'customers/bulk_upload.html', context)


# Download Sample Excel
@login_required(login_url='login')
def download_sample_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Headers
    headers = ['First Name', 'Last Name', 'Email', 'Phone', 'Address', 
               'City', 'State', 'Country', 'Postal Code', 'Company', 'Notes']
    ws.append(headers)
    
    # Sample data
    ws.append(['John', 'Doe', 'john@example.com', '+1234567890', '123 Main St', 
               'New York', 'NY', 'USA', '10001', 'ABC Corp', 'Sample customer'])
    
    # Style headers
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=customer_sample.xlsx'
    return response


# Export to PDF
@login_required(login_url='login')
def export_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4e73df'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    title = Paragraph("Customer Data Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Date
    date_str = datetime.now().strftime('%B %d, %Y')
    date_para = Paragraph(f"Generated on: {date_str}", styles['Normal'])
    elements.append(date_para)
    elements.append(Spacer(1, 20))
    
    # Table data
    customers = Customer.objects.all()
    data = [['Name', 'Email', 'Phone', 'Company', 'City']]
    
    for customer in customers:
        data.append([
            customer.get_full_name(),
            customer.email,
            customer.phone or 'N/A',
            customer.company or 'N/A',
            customer.city or 'N/A'
        ])
    
    # Create table
    table = Table(data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4e73df')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=customers_{datetime.now().strftime("%Y%m%d")}.pdf'
    return response


# User List View
@login_required(login_url='login')
def user_list(request):
    users = User.objects.all().order_by('-date_joined')
    context = {'users': users}
    return render(request, 'customers/user_list.html', context)


# User Add View
@login_required(login_url='login')
def user_add(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            UserProfile.objects.create(user=user)
            messages.success(request, 'User created successfully!')
            return redirect('user_list')
    else:
        form = UserRegistrationForm()
    
    context = {'form': form, 'title': 'Add User'}
    return render(request, 'customers/user_form.html', context)


# User Edit View
@login_required(login_url='login')
def user_edit(request, pk):
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'User updated successfully!')
            return redirect('user_list')
    else:
        form = UserEditForm(instance=user)
    
    context = {'form': form, 'title': 'Edit User', 'user': user}
    return render(request, 'customers/user_form.html', context)


# User Detail View
@login_required(login_url='login')
def user_detail(request, pk):
    user = get_object_or_404(User, pk=pk)
    try:
        profile = user.profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=user)
    
    context = {'user': user, 'profile': profile}
    return render(request, 'customers/user_detail.html', context)


# Edit Profile View
@login_required(login_url='login')
def edit_profile(request):
    user = request.user
    try:
        profile = user.profile
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=user)
    
    if request.method == 'POST':
        user_form = UserEditForm(request.POST, instance=user)
        profile_form = UserProfileForm(request.POST, request.FILES, instance=profile)
        
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('edit_profile')
    else:
        user_form = UserEditForm(instance=user)
        profile_form = UserProfileForm(instance=profile)
    
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'profile': profile
    }
    return render(request, 'customers/edit_profile.html', context)


# Login View
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.first_name}!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'customers/login.html')


# Logout View
@login_required(login_url='login')
def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully!')
    return redirect('login')